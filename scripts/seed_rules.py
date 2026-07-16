#!/usr/bin/env python3
"""
seed_rules.py — Carga reglas de planificación desde Excel a Supabase PostgreSQL.

Hojas:
  - CÓDIGO PLAZO Y DEMORAS EN DÍAS  → tabla lead_time_by_code
  - CAPACIDAD POR PROCESO           → tabla process_capacity

Idempotente: usa ON CONFLICT para upsert en ambas tablas.
  lead_time_by_code : UNIQUE (codigo_plazo, proceso)
  process_capacity  : UNIQUE (proceso)
"""

import sys
import os
import uuid
from datetime import datetime, timezone
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl no instalado. Ejecuta: pip3 install openpyxl")
    sys.exit(1)

try:
    import psycopg2
except ImportError:
    print("ERROR: psycopg2 no instalado. Ejecuta: pip3 install psycopg2-binary")
    sys.exit(1)

# ─── Rutas ────────────────────────────────────────────────────────────────────
SCRIPT_DIR   = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parent
EXCEL_PATH   = PROJECT_ROOT / "data" / "Reglas_Planificación.xlsx"
ENV_PATH     = PROJECT_ROOT / "etp-app" / ".env.local"

# ─── Cargar DATABASE_URL desde .env.local ─────────────────────────────────────
def load_db_url() -> str:
    if not ENV_PATH.exists():
        raise FileNotFoundError(f"No encontré .env.local en {ENV_PATH}")
    with open(ENV_PATH) as f:
        for line in f:
            line = line.strip()
            if line.startswith("DATABASE_URL="):
                url = line.split("=", 1)[1].strip().strip('"').strip("'")
                return url.replace("%21", "!")
    raise RuntimeError("DATABASE_URL no encontrado en .env.local")

# ─── Main ─────────────────────────────────────────────────────────────────────
def main():
    print("=" * 60)
    print("  ETP — Seed de reglas de planificación → Supabase")
    print("=" * 60)

    if not EXCEL_PATH.exists():
        print(f"  ✗ Archivo no encontrado: {EXCEL_PATH}", file=sys.stderr)
        sys.exit(1)

    print(f"\n📂 Leyendo: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(str(EXCEL_PATH), data_only=True)

    # ── Hoja 1: lead_time_by_code ─────────────────────────────────────────────
    ws_lt     = wb["CÓDIGO PLAZO Y DEMORAS EN DÍAS"]
    rows_lt   = list(ws_lt.iter_rows(values_only=True))
    header    = rows_lt[0]
    proc_cols = list(header[2:])          # INGENIERÍA, CORTE, … (11 procesos)

    print(f"  Hoja 'CÓDIGO PLAZO Y DEMORAS EN DÍAS': {len(rows_lt)-1} filas de datos")
    print(f"  Procesos detectados: {[p for p in proc_cols if p]}")

    lead_times = []
    for row in rows_lt[1:]:
        if row[0] is None:
            continue
        codigo_plazo = str(int(row[0]))
        descripcion  = str(row[1]).strip() if row[1] else None
        for col_idx, proc_name in enumerate(proc_cols):
            if proc_name is None:
                continue
            duracion = row[2 + col_idx]
            lead_times.append({
                "id":               str(uuid.uuid4()),
                "codigo_plazo":     codigo_plazo,
                "descripcion_equipo": descripcion,
                "proceso":          str(proc_name).strip(),
                "duracion_dias":    int(duracion) if duracion is not None else 0,
                "now":              datetime.now(timezone.utc),
            })

    print(f"  Registros lead_time preparados: {len(lead_times)}")

    # ── Hoja 2: process_capacity ──────────────────────────────────────────────
    ws_pc   = wb["CAPACIDAD POR PROCESO"]
    rows_pc = list(ws_pc.iter_rows(values_only=True))

    print(f"\n  Hoja 'CAPACIDAD POR PROCESO': {len(rows_pc)-1} filas de datos")

    capacities = []
    for row in rows_pc[1:]:
        if row[0] is None:
            continue
        capacities.append({
            "id":               str(uuid.uuid4()),
            "proceso":          str(row[0]).strip(),
            "orden":            int(row[1]) if row[1] is not None else 0,
            "capacidad_por_dia": int(row[2]) if row[2] is not None else 0,
            "now":              datetime.now(timezone.utc),
        })

    print(f"  Registros process_capacity preparados: {len(capacities)}")
    wb.close()

    # ── Conectar ──────────────────────────────────────────────────────────────
    db_url = load_db_url()
    print(f"\n🔌 Conectando a Supabase PostgreSQL...")
    conn = psycopg2.connect(db_url, sslmode="require")
    conn.autocommit = False
    cur  = conn.cursor()
    print("  Conexión OK")

    # ── Upsert lead_time_by_code ──────────────────────────────────────────────
    # UNIQUE constraint: (codigo_plazo, proceso)
    UPSERT_LT = """
        INSERT INTO lead_time_by_code
            (id, codigo_plazo, descripcion_equipo, proceso, duracion_dias, created_at, updated_at)
        VALUES
            (%(id)s, %(codigo_plazo)s, %(descripcion_equipo)s, %(proceso)s,
             %(duracion_dias)s, %(now)s, %(now)s)
        ON CONFLICT (codigo_plazo, proceso) DO UPDATE SET
            descripcion_equipo = EXCLUDED.descripcion_equipo,
            duracion_dias      = EXCLUDED.duracion_dias,
            updated_at         = EXCLUDED.updated_at
    """

    lt_inserted = lt_updated = lt_errors = 0
    for rec in lead_times:
        try:
            cur.execute("SELECT id FROM lead_time_by_code WHERE codigo_plazo=%s AND proceso=%s",
                        (rec["codigo_plazo"], rec["proceso"]))
            exists = cur.fetchone()
            cur.execute(UPSERT_LT, rec)
            if exists:
                lt_updated += 1
            else:
                lt_inserted += 1
        except Exception as e:
            lt_errors += 1
            print(f"  ✗ lead_time {rec['codigo_plazo']}/{rec['proceso']}: {e}")
            conn.rollback()
    conn.commit()

    # ── Upsert process_capacity ───────────────────────────────────────────────
    # UNIQUE constraint: (proceso)
    UPSERT_PC = """
        INSERT INTO process_capacity
            (id, proceso, orden, capacidad_por_dia, created_at, updated_at)
        VALUES
            (%(id)s, %(proceso)s, %(orden)s, %(capacidad_por_dia)s, %(now)s, %(now)s)
        ON CONFLICT (proceso) DO UPDATE SET
            orden             = EXCLUDED.orden,
            capacidad_por_dia = EXCLUDED.capacidad_por_dia,
            updated_at        = EXCLUDED.updated_at
    """

    pc_inserted = pc_updated = pc_errors = 0
    for rec in capacities:
        try:
            cur.execute("SELECT id FROM process_capacity WHERE proceso=%s", (rec["proceso"],))
            exists = cur.fetchone()
            cur.execute(UPSERT_PC, rec)
            if exists:
                pc_updated += 1
            else:
                pc_inserted += 1
        except Exception as e:
            pc_errors += 1
            print(f"  ✗ process_capacity {rec['proceso']}: {e}")
            conn.rollback()
    conn.commit()

    # ── Verificar conteos finales ─────────────────────────────────────────────
    cur.execute("SELECT COUNT(*) FROM lead_time_by_code")
    total_lt = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM process_capacity")
    total_pc = cur.fetchone()[0]

    cur.close()
    conn.close()

    # ── Resumen ───────────────────────────────────────────────────────────────
    print()
    print("=" * 60)
    print("  RESUMEN")
    print("=" * 60)
    print(f"  Archivo Excel:   {EXCEL_PATH.name}")
    print()
    print(f"  lead_time_by_code:")
    print(f"    Insertados:    {lt_inserted}")
    print(f"    Actualizados:  {lt_updated}")
    print(f"    Errores:       {lt_errors}")
    print(f"    Total en DB:   {total_lt}")
    print()
    print(f"  process_capacity:")
    print(f"    Insertados:    {pc_inserted}")
    print(f"    Actualizados:  {pc_updated}")
    print(f"    Errores:       {pc_errors}")
    print(f"    Total en DB:   {total_pc}")
    print()
    if total_lt >= 187:
        print("  ✅ lead_time_by_code: conteo OK (≥ 187)")
    else:
        print(f"  ⚠  lead_time_by_code: conteo bajo ({total_lt}, esperado ≥ 187)")
    if total_pc >= 11:
        print("  ✅ process_capacity: conteo OK (≥ 11)")
    else:
        print(f"  ⚠  process_capacity: conteo bajo ({total_pc}, esperado 11)")
    print()
    print("  ✅ Refrescar la app en http://localhost:3000")
    print("=" * 60)

if __name__ == "__main__":
    main()
