#!/usr/bin/env python3
"""
seed_sales_planning.py
Importa registros desde 'Programación Equipos.xlsx' a la tabla sales_planning en Supabase.

Hoja usada: Produccion de Equipos
Deduplicación: ot (upsert manual — SELECT+INSERT/UPDATE, ot no es UNIQUE en schema)
"""

import os
import sys
import uuid
from datetime import datetime, timezone
from pathlib import Path

import openpyxl
import psycopg2
import psycopg2.extras

# ─── Rutas ────────────────────────────────────────────────────────────────────
ROOT = Path(__file__).resolve().parent.parent.parent  # .../Proyecto Producción ETP spa
EXCEL_PATH = ROOT / "data" / "Programación Equipos.xlsx"
ENV_PATH   = Path(__file__).resolve().parent.parent / ".env.local"
SHEET_NAME = "Produccion de Equipos"
SEED_USER  = "seed_import"

# ─── Variables de entorno ─────────────────────────────────────────────────────
def load_db_url() -> str:
    if not ENV_PATH.exists():
        raise FileNotFoundError(f"No encontré .env.local en {ENV_PATH}")
    with open(ENV_PATH) as f:
        for line in f:
            line = line.strip()
            if line.startswith("DATABASE_URL="):
                url = line.split("=", 1)[1].strip().strip('"').strip("'")
                return url.replace("%21", "!")
    raise RuntimeError("DATABASE_URL no encontrado en .env.local")

# ─── Helpers de conversión ────────────────────────────────────────────────────
def clean_str(v) -> "str | None":
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None

def to_bool(v) -> bool:
    if isinstance(v, bool):
        return v
    if isinstance(v, str):
        return v.strip().upper() in ("SI", "SÍ", "TRUE", "1", "YES")
    return False

def to_datetime(v) -> "datetime | None":
    if isinstance(v, datetime):
        return v.replace(tzinfo=timezone.utc) if v.tzinfo is None else v
    if isinstance(v, str):
        s = v.strip().split("\n")[0].strip()  # tomar primera línea si hay notas
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
            try:
                return datetime.strptime(s, fmt).replace(tzinfo=timezone.utc)
            except ValueError:
                continue
    return None

# ─── Mapeo de fila Excel → dict ───────────────────────────────────────────────
# Índices (0-based):
#  0  Prioridad        → skip (valor es 'OK'/None, no numérico útil → default 5)
#  1  OT               → ot
#  2  CLTE INTERNO     → clte_interno
#  3  CLIENTE          → cliente
#  4  EQUIPO           → equipo
#  5  MODELO/CAPACIDAD → modelo_capacidad
#  6  CAMION           → camion
#  7  MODELO           → modelo
#  8  VIN              → vin
#  9  LLEGADA          → llegada
# 10  ENTREGA          → SKIP (campo eliminado)
# 11  $ VENTA          → venta
# 12  COLOR EQ.        → color_eq
# 13  OC               → oc
# 14  FACTURA          → factura
# 15  [notas/fecha]    → SKIP (header es datetime, contenido son notas)
# 16  PROXIMO A ENTREGA→ SKIP (campo eliminado)
# 17  COTIZACION       → cotizacion
# 18  CORREO           → correo ("SI" si True, null si False/None)
# 19  PATENTE          → patente
# 20  NEUMATICO REPUESTO→ neumatico_de_repuesto
# 21  N° RECEPCION     → n_recepcion
# 22  COLOR CABINA     → color_cabina

def map_row(row: tuple, row_num: int) -> "dict | None":
    ot = clean_str(row[1])
    if not ot:
        print(f"  ⚠  Fila {row_num}: sin OT → omitida")
        return None

    now = datetime.now(timezone.utc)
    return {
        "id":                   str(uuid.uuid4()),
        "ot":                   ot,
        "clte_interno":         clean_str(row[2]),
        "cliente":              clean_str(row[3]),
        "equipo":               clean_str(row[4]),
        "modelo_capacidad":     clean_str(row[5]),
        "camion":               clean_str(row[6]),
        "modelo":               clean_str(row[7]),
        "vin":                  clean_str(row[8]),
        "llegada":              to_datetime(row[9]),
        # row[10] ENTREGA  → skip
        "venta":                clean_str(row[11]),
        "color_eq":             clean_str(row[12]),
        "oc":                   clean_str(row[13]),
        "factura":              clean_str(row[14]),
        # row[15] notas    → skip
        # row[16] PROXIMO  → skip
        "cotizacion":           to_bool(row[17]),
        "correo":               "SI" if to_bool(row[18]) else None,
        "patente":              clean_str(row[19]),
        "neumatico_de_repuesto": clean_str(row[20]),
        "n_recepcion":          clean_str(row[21]),
        "color_cabina":         clean_str(row[22]),
        "prioridad":            5,   # default; col 0 es marcador 'OK'/None
        "codigo_plazo":         None,
        "created_by":           SEED_USER,
        "updated_by":           SEED_USER,
        "created_at":           now,
        "updated_at":           now,
    }

# ─── SQL ──────────────────────────────────────────────────────────────────────
INSERT_SQL = """
    INSERT INTO sales_planning (
        id, ot, clte_interno, cliente, equipo, modelo_capacidad, camion, modelo,
        vin, llegada, venta, color_eq, oc, factura, cotizacion, correo, patente,
        neumatico_de_repuesto, n_recepcion, color_cabina, prioridad, codigo_plazo,
        created_by, updated_by, created_at, updated_at
    ) VALUES (
        %(id)s, %(ot)s, %(clte_interno)s, %(cliente)s, %(equipo)s,
        %(modelo_capacidad)s, %(camion)s, %(modelo)s, %(vin)s, %(llegada)s,
        %(venta)s, %(color_eq)s, %(oc)s, %(factura)s, %(cotizacion)s,
        %(correo)s, %(patente)s, %(neumatico_de_repuesto)s, %(n_recepcion)s,
        %(color_cabina)s, %(prioridad)s, %(codigo_plazo)s,
        %(created_by)s, %(updated_by)s, %(created_at)s, %(updated_at)s
    )
"""

UPDATE_SQL = """
    UPDATE sales_planning SET
        clte_interno          = %(clte_interno)s,
        cliente               = %(cliente)s,
        equipo                = %(equipo)s,
        modelo_capacidad      = %(modelo_capacidad)s,
        camion                = %(camion)s,
        modelo                = %(modelo)s,
        vin                   = %(vin)s,
        llegada               = %(llegada)s,
        venta                 = %(venta)s,
        color_eq              = %(color_eq)s,
        oc                    = %(oc)s,
        factura               = %(factura)s,
        cotizacion            = %(cotizacion)s,
        correo                = %(correo)s,
        patente               = %(patente)s,
        neumatico_de_repuesto = %(neumatico_de_repuesto)s,
        n_recepcion           = %(n_recepcion)s,
        color_cabina          = %(color_cabina)s,
        prioridad             = %(prioridad)s,
        updated_by            = %(updated_by)s,
        updated_at            = %(updated_at)s
    WHERE ot = %(ot)s
"""

# ─── Main ─────────────────────────────────────────────────────────────────────
def main():
    print("=" * 60)
    print("  ETP — Seed sales_planning desde Excel")
    print("=" * 60)

    # 1. Leer Excel
    print(f"\n📂 Leyendo: {EXCEL_PATH}")
    if not EXCEL_PATH.exists():
        print(f"  ✗ Archivo no encontrado: {EXCEL_PATH}", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(str(EXCEL_PATH), data_only=True)
    ws = wb[SHEET_NAME]
    total_excel_rows = ws.max_row - 1  # sin header
    print(f"  Hoja: '{SHEET_NAME}' | Filas de datos: {total_excel_rows}")

    rows_data = []
    errors = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        try:
            record = map_row(row, i)
            if record:
                rows_data.append(record)
        except Exception as e:
            errors.append((i, str(e)))
            print(f"  ✗ Fila {i} error al mapear: {e}")
    wb.close()

    print(f"  Filas mapeadas válidas: {len(rows_data)}")

    # 2. Conectar a Supabase
    db_url = load_db_url()
    print(f"\n🔌 Conectando a Supabase PostgreSQL...")
    conn = psycopg2.connect(db_url, sslmode="require")
    conn.autocommit = False
    cur = conn.cursor()
    print("  Conexión OK")

    # 3. Cargar OTs existentes
    cur.execute("SELECT ot FROM sales_planning WHERE ot IS NOT NULL")
    existing_ots = {row[0] for row in cur.fetchall()}
    print(f"  OTs ya existentes en DB: {len(existing_ots)}")

    # 4. Upsert
    inserted = 0
    updated  = 0
    skipped  = 0

    for record in rows_data:
        try:
            if record["ot"] in existing_ots:
                record["updated_at"] = datetime.now(timezone.utc)
                cur.execute(UPDATE_SQL, record)
                updated += 1
            else:
                cur.execute(INSERT_SQL, record)
                existing_ots.add(record["ot"])
                inserted += 1
        except Exception as e:
            skipped += 1
            errors.append((record["ot"], str(e)))
            print(f"  ✗ OT {record['ot']}: {e}")
            conn.rollback()
            conn.autocommit = False

    conn.commit()
    cur.close()
    conn.close()

    # 5. Resumen
    print()
    print("=" * 60)
    print("  RESUMEN")
    print("=" * 60)
    print(f"  Hoja Excel:      {SHEET_NAME}")
    print(f"  Filas leídas:    {total_excel_rows}")
    print(f"  Insertados:      {inserted}")
    print(f"  Actualizados:    {updated}")
    print(f"  Omitidos/error:  {skipped + (total_excel_rows - len(rows_data))}")
    if errors:
        print(f"\n  Errores ({len(errors)}):")
        for ref, msg in errors:
            print(f"    - {ref}: {msg}")
    print()
    print("  Columnas mapeadas:")
    mapped = [
        "ot", "clte_interno", "cliente", "equipo", "modelo_capacidad",
        "camion", "modelo", "vin", "llegada", "venta", "color_eq", "oc",
        "factura", "cotizacion", "correo", "patente", "neumatico_de_repuesto",
        "n_recepcion", "color_cabina", "prioridad",
    ]
    for col in mapped:
        print(f"    ✓ {col}")
    print()
    print("  Columnas omitidas (no están en schema):")
    for col in ["entrega", "proximo_a_entrega", "notas_internas"]:
        print(f"    — {col}")
    print()
    print("  ✅ Refrescar la app en http://localhost:3000 para ver los registros.")
    print("=" * 60)

if __name__ == "__main__":
    main()
